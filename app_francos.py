import streamlit as st
import pandas as pd
from ortools.sat.python import cp_model
import calendar
from datetime import datetime
import io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Gestor de Francos AR", page_icon="🗓️", layout="wide")

# --- ESTILOS CSS ---
st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header [data-testid="stHeaderActionElements"] { display: none !important; }
    </style>
    """, unsafe_allow_html=True)

# --- LÓGICA: CALCULAR ARRASTRE ---
def procesar_historial_mes_anterior(df_mes_pasado):
    try:
        columnas_dias = [c for c in df_mes_pasado.columns if str(c).isdigit()]
        columnas_dias.sort(key=int, reverse=True) 
        datos_procesados = []
        for _, fila in df_mes_pasado.iterrows():
            conteo = 0
            for dia in columnas_dias:
                if fila[dia] == 'T': conteo += 1
                else: break
            datos_procesados.append({'Agente': fila['Agente'], 'Tipo': fila['Tipo'], 'Dias_Acumulados': conteo})
        return pd.DataFrame(datos_procesados)
    except Exception as e:
        st.error(f"Error mes anterior: {e}"); return None

# --- LÓGICA DEL OPTIMIZADOR ---
def optimizar_francos(df_empleados, mes_num, anio, reglas_cobertura, dias_activos):
    num_dias = calendar.monthrange(anio, mes_num)[1]
    domingos = [d for d in range(1, num_dias + 1) if calendar.weekday(anio, mes_num, d) == 6]
    cant_francos_objetivo = len(domingos)
    
    model = cp_model.CpModel()
    empleados = df_empleados.to_dict('records')
    num_emp = len(empleados)
    
    x = {} # 1 si Trabaja, 0 si es Franco
    for i in range(num_emp):
        for d in range(1, num_dias + 1):
            x[i, d] = model.NewBoolVar(f'x_{i}_{d}')

    for i, emp in enumerate(empleados):
        tipo = str(emp['Tipo']).strip().capitalize()
        arrastre = int(emp.get('Dias_Acumulados', 0))
        
        # Cantidad de francos: permitimos un pequeño margen de maniobra
        total_trabajo = sum(x[i, d] for d in range(1, num_dias + 1))
        model.Add(total_trabajo >= (num_dias - cant_francos_objetivo - 1))
        model.Add(total_trabajo <= (num_dias - cant_francos_objetivo + 1))

        # --- REGLA DE DESCANSO (Máximo 7 días seguidos de trabajo total) ---
        for d in range(1, num_dias - 6):
            # En cualquier ventana de 8 días, debe haber al menos 1 franco
            model.Add(sum(x[i, d + j] for j in range(8)) <= 7)
        
        # Si el arrastre es crítico (ej: 6 o 7 días), forzamos franco en los primeros 2 días
        if arrastre >= 6:
            model.Add(sum(x[i, d] for d in range(1, 3)) <= 1)

    # --- REGLAS DE COBERTURA (Solo si el check está activo) ---
    for d in range(1, num_dias + 1):
        dia_semana_num = calendar.weekday(anio, mes_num, d)
        nombres_dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        dia_nombre = nombres_dias[dia_semana_num]
        
        # AQUÍ ESTÁ EL TRUCO: Solo entra si el check está tildado
        if dias_activos.get(dia_nombre, False):
            rango = reglas_cobertura.get(dia_nombre, (0, 100))
            min_f = int(num_emp * (rango[0] / 100))
            max_f = int(num_emp * (rango[1] / 100))
            
            # Asegurar que no pedimos más francos de los que hay empleados
            min_f = min(min_f, num_emp - 1)
            
            model.Add(sum(1 - x[i, d] for i in range(num_emp)) >= min_f)
            if max_f > 0:
                model.Add(sum(1 - x[i, d] for i in range(num_emp)) <= max_f)

    # --- FUNCIÓN OBJETIVO (Lo que el sistema 'prefiere' hacer si no hay reglas) ---
    preferencias = []
    for i in range(num_emp):
        for d in range(1, num_dias + 1):
            # Preferir francos en Sábado (5) y Domingo (6)
            if calendar.weekday(anio, mes_num, d) >= 5:
                preferencias.append(1 - x[i, d])
    
    if preferencias:
        model.Maximize(sum(preferencias))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 10.0
    status = solver.Solve(model)

    if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        res = []
        for i, emp in enumerate(empleados):
            # Usar el valor real de i para reconstruir la fila
            fila = {"Agente": emp['Agente'], "Tipo": emp['Tipo'], "Arrastre": int(emp.get('Dias_Acumulados', 0))}
            for d in range(1, num_dias + 1):
                fila[f"{d}"] = "T" if solver.Value(x[i, d]) == 1 else "F"
            res.append(fila)
        return pd.DataFrame(res)
    return None
    num_dias = calendar.monthrange(anio, mes_num)[1]
    domingos = [d for d in range(1, num_dias + 1) if calendar.weekday(anio, mes_num, d) == 6]
    
    # Cantidad base de francos (normalmente 4 o 5)
    cant_francos_objetivo = len(domingos)
    
    model = cp_model.CpModel()
    empleados = df_empleados.to_dict('records')
    num_emp = len(empleados)
    
    x = {} # 1 si Trabaja, 0 si es Franco
    for i in range(num_emp):
        for d in range(1, num_dias + 1):
            x[i, d] = model.NewBoolVar(f'x_{i}_{d}')

    for i, emp in enumerate(empleados):
        tipo = str(emp['Tipo']).strip().capitalize()
        arrastre = int(emp.get('Dias_Acumulados', 0))
        
        # --- CAMBIO CLAVE: Flexibilidad en la cantidad de francos ---
        # En lugar de == exacto, permitimos un margen de +/- 1 día si el modelo está muy apretado
        total_trabajo = sum(x[i, d] for d in range(1, num_dias + 1))
        model.Add(total_trabajo >= (num_dias - cant_francos_objetivo - 1))
        model.Add(total_trabajo <= (num_dias - cant_francos_objetivo + 1))

        if tipo == 'Tercerizado':
            # Si el arrastre es muy alto, obligamos franco el día 1 o 2
            if arrastre >= 7:
                model.Add(x[i, 1] == 0)
            # Regla general: máximo 7 días seguidos de trabajo
            for d in range(1, num_dias - 6):
                model.Add(sum(x[i, d + j] for j in range(8)) <= 7)

        elif tipo == 'Propio':
            # Regla 6x2: Relajamos un poco la ventana para que sea factible
            # En cualquier ventana de 9 días, debe haber al menos 2 francos
            for d in range(1, num_dias - 8):
                model.Add(sum(1 - x[i, d + j] for j in range(9)) >= 2)
            
            # Si arrastra mucho trabajo, forzar franco en los primeros días
            if arrastre >= 6:
                model.Add(sum(x[i, d] for d in range(1, 3)) <= 1)

    # --- REGLAS DE COBERTURA (Solo si están activas) ---
    for d in range(1, num_dias + 1):
        dia_semana_num = calendar.weekday(anio, mes_num, d)
        nombres_dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        dia_nombre = nombres_dias[dia_semana_num]
        
        if dias_activos.get(dia_nombre, False):
            rango = reglas_cobertura.get(dia_nombre, (0, 100))
            min_f = int(num_emp * (rango[0] / 100))
            # Asegurar que siempre quede al menos 1 persona trabajando
            min_f = min(min_f, num_emp - 1) 
            
            model.Add(sum(1 - x[i, d] for i in range(num_emp)) >= min_f)
            
            if rango[1] > 0:
                max_f = int(num_emp * (rango[1] / 100))
                model.Add(sum(1 - x[i, d] for i in range(num_emp)) <= max_f)

    # --- FUNCIÓN OBJETIVO: Intentar que los francos caigan en fin de semana si es posible ---
    # Esto ayuda a que el sistema "elija" mejor cuando no hay restricciones tildadas
    pref_francos = []
    for i in range(num_emp):
        for d in range(1, num_dias + 1):
            if calendar.weekday(anio, mes_num, d) >= 5: # Sábado o Domingo
                pref_francos.append(1 - x[i, d])
    model.Maximize(sum(pref_francos))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 10.0
    status = solver.Solve(model)

    if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        res = []
        for i, emp in enumerate(empleados):
            fila = {"Agente": emp['Agente'], "Tipo": emp['Tipo'], "Arrastre": int(emp.get('Dias_Acumulados', 0))}
            for d in range(1, num_dias + 1):
                fila[f"{d}"] = "T" if solver.Value(x[i, d]) == 1 else "F"
            res.append(fila)
        return pd.DataFrame(res)
    return None

def exportar_excel_formateado(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Planificacion')
        ws = writer.sheets['Planificacion']
        f_fill, f_font = PatternFill(start_color="FFC7CE", fill_type="solid"), Font(color="9C0006", bold=True)
        t_fill, t_font = PatternFill(start_color="C6EFCE", fill_type="solid"), Font(color="006100")
        for row in range(2, ws.max_row + 1):
            for col in range(4, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value == "F": cell.fill, cell.font = f_fill, f_font
                elif cell.value == "T": cell.fill, cell.font = t_fill, t_font
                cell.alignment = Alignment(horizontal='center')
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20 if col == 1 else 4
    return output.getvalue()

# --- UI ---
st.title("🗓️ Gestor de Francos")
with st.sidebar:
    st.header("⚙️ Configuración")
    meses_es = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    mes_nombre = st.selectbox("Mes a planificar", meses_es, index=(datetime.now().month)%12)
    mes_num = meses_es.index(mes_nombre) + 1
    anio = st.number_input("Año", value=datetime.now().year, step=1)
    
    st.divider()
    st.subheader("📊 % Francos por Día")
    st.caption("Tildá el día para activar su restricción de cobertura.")
    
    reglas_cobertura = {}
    dias_activos = {}
    dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    
    # --- UI de Sliders con Checkbox corregido ---
    for dia in dias_semana:
        def_val = (30, 50) if dia == "Domingo" else (10, 20) if dia == "Sábado" else (5, 15)
        
        col1, col2 = st.columns([1, 4])
        with col1:
            # Le ponemos un label real pero lo ocultamos para que no de error
            dias_activos[dia] = st.checkbox(
                f"Activar {dia}", 
                value=False, 
                key=f"check_{dia}", 
                label_visibility="collapsed"
            )
        with col2:
            reglas_cobertura[dia] = st.slider(
                f"{dia}", 
                0, 100, 
                def_val, 
                disabled=not dias_activos[dia]
            )

    st.divider()
    modo_carga = st.radio("Método de arrastre:", ["Usar planilla mes pasado", "Manual (personal.xlsx)"])

file = st.file_uploader("Subir archivo Excel", type=["xlsx"])
if file:
    df_input = procesar_historial_mes_anterior(pd.read_excel(file)) if modo_carga == "Usar planilla mes pasado" else pd.read_excel(file)
    if st.button("🚀 Generar Planificación"):
        res_df = optimizar_francos(df_input, mes_num, anio, reglas_cobertura, dias_activos)
        if res_df is not None:
            st.dataframe(res_df, use_container_width=True)
            st.download_button("📥 Descargar Excel", exportar_excel_formateado(res_df), f"Planificacion_{mes_nombre}.xlsx")
        else:
            st.error("❌ Sin solución. Probá destildando más días o bajando los porcentajes.")